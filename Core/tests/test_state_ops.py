"""
Tests for state operation requirements:
  CHAT-03 — Undo reverts the last command and syncs chart state
  SESS-01  — Authenticated user can list their conversations
  SESS-02  — Conversation state can be resumed; access denied for other users
  AUTH-03  — Conversation list is scoped to the authenticated user
  FILE-02  — TTL-based cleanup removes expired uploaded files (stubs — plan 03)
"""
import json
import io
import os
import pytest
import openpyxl
from unittest.mock import patch
from uuid import uuid4


# ---------------------------------------------------------------------------
# Helpers (unique emails via uuid4 to avoid collisions with other test suites)
# ---------------------------------------------------------------------------

def _make_authenticated_client(client, email=None):
    """Register + login a test user and return the JWT token."""
    if email is None:
        email = f"statetest_{uuid4().hex[:8]}@example.com"
    client.post('/auth/register', json={'email': email, 'password': 'Password1!'})
    resp = client.post('/auth/login', json={'email': email, 'password': 'Password1!'})
    data = resp.get_json()
    return data.get('token') or data.get('access_token') or ''


def _upload_test_xlsx(client, token):
    """Create a minimal .xlsx in memory, upload it, return session_id."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Name'
    ws['B1'] = 'Value'
    ws['A2'] = 'Alice'
    ws['B2'] = 100
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = client.post(
        '/excel/upload',
        data={'file': (buf, 'test.xlsx')},
        headers={'Authorization': f'Bearer {token}'},
        content_type='multipart/form-data'
    )
    assert resp.status_code == 200, f"Upload failed: {resp.data}"
    return resp.get_json()['session_id']


def _post_transform_with_mock(client, token, session_id, prompt, mock_return):
    """
    POST /excel/transform with LLMService mocked for the full SSE lifecycle.

    Flask's test client evaluates the SSE generator lazily when resp.data is
    first accessed.  The mock must remain active through that access, so
    patcher.stop() is called only AFTER resp.data is consumed.
    """
    patcher = patch(
        'app.routes.excel.LLMService.generate_transformation_code',
        return_value=mock_return
    )
    patcher.start()
    try:
        resp = client.post(
            '/excel/transform',
            data={'session_id': str(session_id), 'prompt': prompt},
            headers={'Authorization': f'Bearer {token}'},
        )
        # Force full stream consumption while mock is still active
        raw_data = resp.data
    finally:
        patcher.stop()
    return resp, raw_data


# ---------------------------------------------------------------------------
# CHAT-03: Undo tests
# ---------------------------------------------------------------------------

def test_undo_reverts_last_command(client):
    """CHAT-03 — undo after a DATA_MUTATION reverts the grid to its pre-command state."""
    token = _make_authenticated_client(client)
    session_id = _upload_test_xlsx(client, token)

    # Apply a transform that doubles B2 from 100 -> 200
    mock_return = {
        'code': "df['Value'] = df['Value'] * 2",
        'explanation': 'Doubled values.',
        'intent': 'DATA_MUTATION'
    }
    _post_transform_with_mock(client, token, session_id, 'double all values', mock_return)

    # Undo
    resp = client.post(
        '/excel/undo',
        data={'session_id': str(session_id)},
        headers={'Authorization': f'Bearer {token}'},
    )
    data = resp.get_json()

    assert resp.status_code == 200, f"Undo returned non-200: {data}"
    assert data.get('status') == 'success'
    # Value should be reverted to 100
    rows = data['data']['rows']
    assert rows[0]['Value'] == 100, f"Expected Value=100 after undo, got: {rows[0]}"


def test_undo_no_commands(client):
    """CHAT-03 — undo on a session with no commands returns 200 with undone=False."""
    token = _make_authenticated_client(client)
    session_id = _upload_test_xlsx(client, token)

    resp = client.post(
        '/excel/undo',
        data={'session_id': str(session_id)},
        headers={'Authorization': f'Bearer {token}'},
    )
    data = resp.get_json()

    assert resp.status_code == 200, f"Undo (empty) returned non-200: {data}"
    assert data.get('undone') == False, f"Expected undone=False, got: {data.get('undone')}"


def test_undo_includes_chart_fields(client):
    """CHAT-03 — undo response always contains chart_data and has_chart keys."""
    token = _make_authenticated_client(client)
    session_id = _upload_test_xlsx(client, token)

    resp = client.post(
        '/excel/undo',
        data={'session_id': str(session_id)},
        headers={'Authorization': f'Bearer {token}'},
    )
    data = resp.get_json()

    assert resp.status_code == 200, f"Undo returned non-200: {data}"
    assert 'chart_data' in data, f"'chart_data' key missing from undo response: {data}"
    assert 'has_chart' in data, f"'has_chart' key missing from undo response: {data}"


# ---------------------------------------------------------------------------
# SESS-01 / AUTH-03: Conversation list
# ---------------------------------------------------------------------------

def test_get_conversations_returns_list(client):
    """SESS-01 / AUTH-03 — authenticated user gets their conversation list (len >= 1)."""
    token = _make_authenticated_client(client)
    _upload_test_xlsx(client, token)

    resp = client.get(
        '/excel/conversations',
        headers={'Authorization': f'Bearer {token}'},
    )
    data = resp.get_json()

    assert resp.status_code == 200, f"GET /conversations returned non-200: {data}"
    assert isinstance(data.get('conversations'), list), f"'conversations' is not a list: {data}"
    assert len(data['conversations']) >= 1, f"Expected >= 1 conversation, got: {data['conversations']}"


# ---------------------------------------------------------------------------
# SESS-02 / CHAT-04: Conversation resume
# ---------------------------------------------------------------------------

def test_conversation_resume_returns_state(client):
    """SESS-02 / CHAT-04 — GET /conversation/<id> returns messages and grid after a transform."""
    token = _make_authenticated_client(client)
    session_id = _upload_test_xlsx(client, token)

    # Apply one transform to populate the message history
    mock_return = {
        'code': "df['Value'] = df['Value'] + 1",
        'explanation': 'Incremented values by 1.',
        'intent': 'DATA_MUTATION'
    }
    _post_transform_with_mock(client, token, session_id, 'increment values', mock_return)

    resp = client.get(
        f'/excel/conversation/{session_id}',
        headers={'Authorization': f'Bearer {token}'},
    )
    data = resp.get_json()

    assert resp.status_code == 200, f"GET /conversation/{session_id} returned non-200: {data}"
    assert 'messages' in data['data'], f"'messages' missing from response data: {data['data']}"
    assert len(data['data']['messages']) >= 2, (
        f"Expected >= 2 messages (user + assistant), got: {data['data']['messages']}"
    )
    assert 'grid' in data['data'], f"'grid' missing from response data: {data['data']}"


def test_conversation_access_denied(client):
    """SESS-02 — a second user cannot access another user's conversation (403)."""
    # Owner uploads a file
    owner_token = _make_authenticated_client(client)
    session_id = _upload_test_xlsx(client, owner_token)

    # Attacker registers and logs in
    attacker_email = f"attacker_{uuid4().hex[:8]}@example.com"
    attacker_token = _make_authenticated_client(client, email=attacker_email)

    # Attacker tries to access owner's conversation
    resp = client.get(
        f'/excel/conversation/{session_id}',
        headers={'Authorization': f'Bearer {attacker_token}'},
    )
    assert resp.status_code == 403, (
        f"Expected 403 for cross-user access, got {resp.status_code}: {resp.get_json()}"
    )


# ---------------------------------------------------------------------------
# FILE-02: TTL cleanup (expires_at column + jobs.py — plan 03)
# ---------------------------------------------------------------------------

def test_ttl_cleanup_removes_file(app):
    """FILE-02: cleanup_expired_files removes disk file + deactivates record when expires_at has passed."""
    import tempfile
    from datetime import datetime, timedelta
    from config.database import SessionLocal
    from app.models.conversation import Conversation
    from app.jobs import cleanup_expired_files

    # Create a real temp file to simulate an uploaded Excel
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        temp_path = f.name

    db = SessionLocal()
    try:
        conv = Conversation(
            user_id=999,  # Fake user ID — no FK constraint in SQLite with test schema
            file_path=temp_path,
            filename='expired.xlsx',
            expires_at=datetime.utcnow() - timedelta(seconds=1)  # Already expired
        )
        db.add(conv)
        db.commit()
        conv_id = conv.id
    finally:
        db.close()

    cleanup_expired_files()

    db = SessionLocal()
    try:
        refreshed = db.query(Conversation).filter_by(id=conv_id).first()
        assert refreshed.is_active is False, "Conversation should be deactivated"
        assert not os.path.exists(temp_path), "File should be removed from disk"
    finally:
        db.close()
        # Cleanup in case test fails
        if os.path.exists(temp_path):
            os.remove(temp_path)


def test_ttl_cleanup_skips_fresh(app):
    """FILE-02: cleanup_expired_files does not touch conversations with future expires_at."""
    import tempfile
    from datetime import datetime, timedelta
    from config.database import SessionLocal
    from app.models.conversation import Conversation
    from app.jobs import cleanup_expired_files

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        temp_path = f.name

    db = SessionLocal()
    try:
        conv = Conversation(
            user_id=998,
            file_path=temp_path,
            filename='fresh.xlsx',
            expires_at=datetime.utcnow() + timedelta(days=7)  # Future
        )
        db.add(conv)
        db.commit()
        conv_id = conv.id
    finally:
        db.close()

    cleanup_expired_files()

    db = SessionLocal()
    try:
        refreshed = db.query(Conversation).filter_by(id=conv_id).first()
        assert refreshed.is_active is True, "Fresh conversation must remain active"
        assert os.path.exists(temp_path), "Fresh file must not be deleted"
    finally:
        db.close()
        if os.path.exists(temp_path):
            os.remove(temp_path)


def test_ttl_cleanup_missing_file(app):
    """FILE-02: cleanup_expired_files does not crash when file_path does not exist on disk."""
    import tempfile
    from datetime import datetime, timedelta
    from config.database import SessionLocal
    from app.models.conversation import Conversation
    from app.jobs import cleanup_expired_files

    ghost_path = os.path.join(tempfile.gettempdir(), 'does_not_exist_datamind_9999.xlsx')

    db = SessionLocal()
    try:
        conv = Conversation(
            user_id=997,
            file_path=ghost_path,
            filename='ghost.xlsx',
            expires_at=datetime.utcnow() - timedelta(seconds=1)
        )
        db.add(conv)
        db.commit()
        conv_id = conv.id
    finally:
        db.close()

    # Must not raise
    cleanup_expired_files()

    db = SessionLocal()
    try:
        refreshed = db.query(Conversation).filter_by(id=conv_id).first()
        assert refreshed.is_active is False, "Record should be deactivated even when file is missing"
    finally:
        db.close()
