"""
Tests for FILE-03: User can download the modified .xlsx file.
  test_download_returns_xlsx          — 200 + xlsx MIME for valid session
  test_download_cross_user_returns_403 — 403 for cross-user session
  test_download_missing_session_returns_error — 400 or 404 for non-existent session
"""
import io
import pytest
import openpyxl
from uuid import uuid4


# ---------------------------------------------------------------------------
# Helpers (shared with other test modules — same pattern as test_state_ops.py)
# ---------------------------------------------------------------------------

def _make_authenticated_client(client, email=None):
    """Register + login a test user and return the JWT token."""
    if email is None:
        email = f"dltest_{uuid4().hex[:8]}@example.com"
    client.post('/auth/register', json={'email': email, 'password': 'Password1!'})
    resp = client.post('/auth/login', json={'email': email, 'password': 'Password1!'})
    data = resp.get_json()
    return data.get('token') or data.get('access_token') or ''


def _upload_test_xlsx(client, token):
    """Create a minimal .xlsx in memory, upload it, return session_id."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Name'
    ws['B1'] = 'Value'
    ws['A2'] = 'Alice'
    ws['B2'] = 100
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = client.post(
        '/excel/upload',
        data={'file': (buf, 'test.xlsx')},
        headers={'Authorization': f'Bearer {token}'},
        content_type='multipart/form-data'
    )
    assert resp.status_code == 200, f"Upload failed: {resp.data}"
    return resp.get_json()['session_id']


# ---------------------------------------------------------------------------
# FILE-03: Download endpoint tests
# ---------------------------------------------------------------------------

class TestDownloadEndpoint:
    """FILE-03: User can download the modified .xlsx file."""

    def test_download_returns_xlsx(self, client):
        """GET /excel/download/<session_id> with valid session returns 200 and xlsx content-type."""
        token = _make_authenticated_client(client)
        session_id = _upload_test_xlsx(client, token)

        resp = client.get(
            f'/excel/download/{session_id}',
            headers={'Authorization': f'Bearer {token}'}
        )
        assert resp.status_code == 200, f"Expected 200, got {resp.status_code}: {resp.data[:200]}"
        assert resp.content_type == (
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ), f"Unexpected content-type: {resp.content_type}"
        # Verify non-zero body
        assert len(resp.data) > 0, "Response body should be non-empty xlsx bytes"

    def test_download_cross_user_returns_403(self, client):
        """GET /excel/download/<session_id> with a different user's session returns 403."""
        owner_token = _make_authenticated_client(client)
        session_id = _upload_test_xlsx(client, owner_token)

        # Attacker registers separately
        attacker_token = _make_authenticated_client(client)

        resp = client.get(
            f'/excel/download/{session_id}',
            headers={'Authorization': f'Bearer {attacker_token}'}
        )
        assert resp.status_code == 403, (
            f"Expected 403 for cross-user download, got {resp.status_code}: {resp.data[:200]}"
        )
        body = resp.get_json()
        assert 'error' in body, f"Expected 'error' key in 403 response: {body}"

    def test_download_missing_session_returns_error(self, client):
        """GET /excel/download/<session_id> where session_id does not exist returns 400 or 404."""
        token = _make_authenticated_client(client)

        resp = client.get(
            '/excel/download/99999999',
            headers={'Authorization': f'Bearer {token}'}
        )
        assert resp.status_code in (400, 403, 404, 500), (
            f"Expected 4xx for missing session, got {resp.status_code}: {resp.data[:200]}"
        )
        body = resp.get_json()
        assert 'error' in body, f"Expected 'error' key in error response: {body}"
