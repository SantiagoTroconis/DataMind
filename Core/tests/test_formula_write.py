"""
Tests for EDIT-04: apply_formula_write writes formula strings to .xlsx disk via openpyxl.
"""
import os
import tempfile
import pytest
import openpyxl
import pandas as pd


def test_formula_written_to_disk():
    """After apply_formula_write, .xlsx file on disk contains formula string in target cell."""
    from app.services.code_execution_service import CodeExecutionService

    # Create a temp .xlsx with numeric data in C1:C9
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp_path = tmp.name

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(1, 10):
            ws[f'C{i}'] = i  # C1=1, C2=2, ..., C9=9
        wb.save(tmp_path)

        # Apply formula write
        CodeExecutionService.apply_formula_write(tmp_path, [
            {"cell": "C10", "formula": "=SUM(C1:C9)"}
        ])

        # Re-read with data_only=False to get formula strings
        wb2 = openpyxl.load_workbook(tmp_path, data_only=False)
        ws2 = wb2.active
        assert ws2['C10'].value == '=SUM(C1:C9)', (
            f"Expected '=SUM(C1:C9)' in C10, got {ws2['C10'].value!r}"
        )
    finally:
        os.unlink(tmp_path)
