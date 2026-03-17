"""
Tests for EDIT-03 and CHAT-02: SSE /excel/transform endpoint.

These tests mock the LLM so no real API calls are made.
"""
import json
import io
import pytest
import openpyxl
from unittest.mock import patch


def _parse_sse_frames(raw: bytes) -> list:
    """Parse SSE byte stream into a list of {'event': str, 'data': dict}."""
    text = raw.decode('utf-8')
    frames = []
    for block in text.split('\n\n'):
        block = block.strip()
        if not block:
            continue
        event = None
        data = None
        for line in block.splitlines():
            if line.startswith('event:'):
                event = line[len('event:'):].strip()
            elif line.startswith('data:'):
                data = json.loads(line[len('data:'):].strip())
        if data is not None:
            frames.append({'event': event or 'message', 'data': data})
    return frames


def _make_authenticated_client(client):
    """Register + login a test user and return token."""
    client.post('/auth/register', json={'email': 'ssetest@example.com', 'password': 'Password1!'})
    resp = client.post('/auth/login', json={'email': 'ssetest@example.com', 'password': 'Password1!'})
    data = resp.get_json()
    token = data.get('token') or data.get('access_token') or ''
    return token


def _upload_test_xlsx(client, token):
    """Upload a minimal .xlsx and return session_id."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Name'
    ws['B1'] = 'Value'
    ws['A2'] = 'Alice'
    ws['B2'] = 100
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = client.post(
        '/excel/upload',
        data={'file': (buf, 'test.xlsx')},
        headers={'Authorization': f'Bearer {token}'},
        content_type='multipart/form-data'
    )
    assert resp.status_code == 200, f"Upload failed: {resp.data}"
    return resp.get_json()['session_id']


def _post_transform_with_mock(client, token, session_id, prompt, mock_return, extra_data=None):
    """
    POST to /excel/transform with LLMService mocked for the entire lifecycle
    including lazy stream consumption.

    Flask's test client for streaming responses evaluates the generator lazily
    when resp.data is first accessed. The mock must remain active through that
    access, so patcher.stop() is called only after resp.data is consumed.
    """
    patcher = patch(
        'app.routes.excel.LLMService.generate_transformation_code',
        return_value=mock_return
    )
    patcher.start()
    try:
        payload = {
            'session_id': str(session_id),
            'prompt': prompt,
        }
        if extra_data:
            payload.update(extra_data)

        resp = client.post(
            '/excel/transform',
            data=payload,
            headers={'Authorization': f'Bearer {token}'},
        )
        # Force full stream consumption while mock is still active
        raw_data = resp.data
    finally:
        patcher.stop()
    return resp, raw_data


def test_transform_sse_events(client):
    """POST /excel/transform yields SSE events: progress(Interpretando), progress(Ejecutando), done."""
    token = _make_authenticated_client(client)
    session_id = _upload_test_xlsx(client, token)

    mock_code_data = {
        'code': "df['Value'] = df['Value'] * 2",
        'explanation': 'Doubled the values.',
        'intent': 'DATA_MUTATION'
    }

    resp, raw_data = _post_transform_with_mock(client, token, session_id, 'double all values', mock_code_data)

    assert resp.status_code == 200
    assert 'text/event-stream' in resp.content_type

    frames = _parse_sse_frames(raw_data)
    events = [f['event'] for f in frames]

    assert 'progress' in events, f"No 'progress' events found. Events: {events}"
    assert 'done' in events, f"No 'done' event found. Events: {events}"

    progress_frames = [f for f in frames if f['event'] == 'progress']
    steps = [f['data'].get('step', '') for f in progress_frames]
    assert any('Interpretando' in s for s in steps), f"Expected 'Interpretando' step, got: {steps}"
    assert any('Ejecutando' in s for s in steps), f"Expected 'Ejecutando' step, got: {steps}"


def test_explanation_in_done_event(client):
    """done SSE event payload contains 'explanation' key with non-empty string."""
    token = _make_authenticated_client(client)
    session_id = _upload_test_xlsx(client, token)

    mock_code_data = {
        'code': "df['Value'] = df['Value'] + 1",
        'explanation': 'Incremented values by 1.',
        'intent': 'DATA_MUTATION'
    }

    resp, raw_data = _post_transform_with_mock(client, token, session_id, 'increment values', mock_code_data)

    assert resp.status_code == 200
    frames = _parse_sse_frames(raw_data)
    done_frames = [f for f in frames if f['event'] == 'done']
    assert len(done_frames) == 1, f"Expected exactly one done event, got: {done_frames}"

    payload = done_frames[0]['data']
    assert 'explanation' in payload, f"'explanation' key missing from done payload: {payload}"
    assert payload['explanation'], f"'explanation' is empty in done payload: {payload}"


def test_selected_range_restriction_for_non_chart_prompt(client):
    """Selected range should only be accepted for VISUAL_UPDATE prompts."""
    token = _make_authenticated_client(client)
    session_id = _upload_test_xlsx(client, token)

    mock_code_data = {
        'code': "df['Value'] = df['Value'] + 1",
        'explanation': 'Incremented values by 1.',
        'intent': 'DATA_MUTATION'
    }
    selected_range = {
        'rangeLabel': 'A2:B2',
        'startCell': 'A2',
        'endCell': 'B2',
        'rowCount': 1,
        'columnCount': 2,
        'cellCount': 2,
        'columns': ['Name', 'Value'],
        'rows': [{'Name': 'Alice', 'Value': 100}],
    }

    resp, raw_data = _post_transform_with_mock(
        client,
        token,
        session_id,
        'increment values',
        mock_code_data,
        extra_data={'selected_range': json.dumps(selected_range)},
    )

    assert resp.status_code == 200
    frames = _parse_sse_frames(raw_data)
    error_frames = [f for f in frames if f['event'] == 'error']
    assert error_frames, f"Expected error SSE frame, got: {frames}"
    assert 'selección de celdas solo aplica para crear gráficas' in error_frames[0]['data'].get('error', '').lower()


def test_visual_update_uses_selected_range_scope(client):
    """VISUAL_UPDATE should build chart data using only the selected range rows/columns."""
    token = _make_authenticated_client(client)
    session_id = _upload_test_xlsx(client, token)

    mock_code_data = {
        'code': "fig = px.bar(df, x='Name', y='Value')",
        'explanation': 'Generated chart from selected cells.',
        'intent': 'VISUAL_UPDATE'
    }
    selected_range = {
        'rangeLabel': 'A2:B2',
        'startCell': 'A2',
        'endCell': 'B2',
        'rowCount': 1,
        'columnCount': 2,
        'cellCount': 2,
        'columns': ['Name', 'Value'],
        'rows': [{'Name': 'Alice', 'Value': 100}],
    }

    with patch('app.routes.excel.CodeExecutionService.execute_chart_generation', return_value='{"data": [], "layout": {}}') as chart_exec_mock:
        resp, raw_data = _post_transform_with_mock(
            client,
            token,
            session_id,
            'create chart',
            mock_code_data,
            extra_data={'selected_range': json.dumps(selected_range)},
        )

    assert resp.status_code == 200
    assert chart_exec_mock.called, "Chart execution should be called"
    scoped_df = chart_exec_mock.call_args[0][0]
    assert list(scoped_df.columns) == ['Name', 'Value']
    assert len(scoped_df) == 1

    frames = _parse_sse_frames(raw_data)
    done_frames = [f for f in frames if f['event'] == 'done']
    assert done_frames, f"Expected done event, got: {frames}"
